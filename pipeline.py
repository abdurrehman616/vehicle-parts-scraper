"""
Main pipeline — orchestrates scraping, parsing, and Excel output.

Usage:
    python3.11 pipeline.py

Produces:
    products.xlsx  with three sheets:
      • Airlift  (40 rows)
      • Timbren  (40 rows)
      • All Products (80 rows)
"""

import asyncio
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper_airlift import scrape_airlift
from scraper_timbren import scrape_timbren
from parser import parse_product

OUTPUT_FILE = Path(__file__).parent / "products.xlsx"

# Column order for the output spreadsheet
COLUMNS = [
    "source",
    "sku",
    "title",
    "price",
    "year_start",
    "year_end",
    "make",
    "model",
    "submodel",
    "product_family",
    "product_type",
    "capacity",
    "position",
    "modifier",
]

COLUMN_LABELS = {
    "source":         "Source",
    "sku":            "SKU",
    "title":          "Product Title",
    "price":          "Price",
    "year_start":     "Year Start",
    "year_end":       "Year End",
    "make":           "Make",
    "model":          "Model",
    "submodel":       "Sub-Model",
    "product_family": "Product Family",
    "product_type":   "Product Type",
    "capacity":       "Capacity",
    "position":       "Position",
    "modifier":       "Modifier",
}

# Brand colours
AIRLIFT_HEADER_BG = "D9182B"   # Air Lift red
TIMBREN_HEADER_BG = "1F4E79"   # Timbren navy
ALL_HEADER_BG     = "2E4057"   # dark slate

HEADER_FG = "FFFFFF"


# ---------------------------------------------------------------------------
# Scrape
# ---------------------------------------------------------------------------

def run_scraping() -> tuple[list[dict], list[dict]]:
    print("=" * 60)
    print("STEP 1  Scraping product data")
    print("=" * 60)

    airlift_raw = asyncio.run(scrape_airlift(limit=40))
    timbren_raw = scrape_timbren(limit=40)

    return airlift_raw, timbren_raw


# ---------------------------------------------------------------------------
# Parse
# ---------------------------------------------------------------------------

def run_parsing(
    airlift_raw: list[dict], timbren_raw: list[dict]
) -> tuple[list[dict], list[dict]]:
    print("\n" + "=" * 60)
    print("STEP 2  Parsing product titles")
    print("=" * 60)

    airlift_parsed = [parse_product(r) for r in airlift_raw]
    timbren_parsed = [parse_product(r) for r in timbren_raw]

    print(f"  Airlift  parsed : {len(airlift_parsed)} records")
    print(f"  Timbren  parsed : {len(timbren_parsed)} records")

    return airlift_parsed, timbren_parsed


# ---------------------------------------------------------------------------
# Build DataFrames
# ---------------------------------------------------------------------------

def build_dataframes(
    airlift: list[dict], timbren: list[dict]
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    def to_df(records: list[dict]) -> pd.DataFrame:
        df = pd.DataFrame(records)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df[COLUMNS]
        # Coerce every column to string so Excel stores text, not floats.
        # This prevents pandas from inferring numeric types for year columns.
        for col in df.columns:
            df[col] = df[col].fillna("").astype(str).replace({"nan": "", "None": ""})
        return df

    df_airlift = to_df(airlift)
    df_timbren = to_df(timbren)
    df_all     = pd.concat([df_airlift, df_timbren], ignore_index=True)

    return df_airlift, df_timbren, df_all


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_sheet(ws, header_bg: str, freeze: bool = True) -> None:
    """Apply header styling, column widths, and row alternation to a sheet."""
    header_fill = PatternFill("solid", fgColor=header_bg)
    alt_fill    = PatternFill("solid", fgColor="F5F5F5")

    max_col = ws.max_column
    max_row = ws.max_row

    # Header row
    for cell in ws[1]:
        cell.font      = Font(bold=True, color=HEADER_FG, size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()

    # Data rows
    for row_idx in range(2, max_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border    = _thin_border()
            cell.font      = Font(size=9)

    # Column widths
    col_widths = {
        "A": 10,  # source
        "B": 14,  # sku
        "C": 50,  # title
        "D": 10,  # price
        "E": 11,  # year_start
        "F": 10,  # year_end
        "G": 12,  # make
        "H": 16,  # model
        "I": 20,  # submodel
        "J": 28,  # product_family
        "K": 28,  # product_type
        "L": 12,  # capacity
        "M": 14,  # position
        "N": 28,  # modifier
    }
    for col_letter, width in col_widths.items():
        if col_letter in [get_column_letter(i) for i in range(1, max_col + 1)]:
            ws.column_dimensions[col_letter].width = width

    # Row heights
    ws.row_dimensions[1].height = 28
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 15

    if freeze:
        ws.freeze_panes = "A2"


def write_excel(
    df_airlift: pd.DataFrame,
    df_timbren: pd.DataFrame,
    df_all:     pd.DataFrame,
    path:       Path,
) -> None:
    print("\n" + "=" * 60)
    print("STEP 3  Writing Excel output")
    print("=" * 60)

    human_labels = [COLUMN_LABELS.get(c, c) for c in COLUMNS]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_airlift.to_excel(writer, sheet_name="Airlift",      index=False, header=human_labels)
        df_timbren.to_excel(writer, sheet_name="Timbren",      index=False, header=human_labels)
        df_all    .to_excel(writer, sheet_name="All Products",  index=False, header=human_labels)

    wb = load_workbook(path)
    _style_sheet(wb["Airlift"],      AIRLIFT_HEADER_BG)
    _style_sheet(wb["Timbren"],      TIMBREN_HEADER_BG)
    _style_sheet(wb["All Products"], ALL_HEADER_BG)
    wb.save(path)

    print(f"  Written → {path}")
    print(f"  Airlift  : {len(df_airlift)} rows")
    print(f"  Timbren  : {len(df_timbren)} rows")
    print(f"  All      : {len(df_all)} rows")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    airlift_raw, timbren_raw = run_scraping()

    if not airlift_raw:
        print("ERROR: No Airlift products collected. Aborting.", file=sys.stderr)
        sys.exit(1)
    if not timbren_raw:
        print("ERROR: No Timbren products collected. Aborting.", file=sys.stderr)
        sys.exit(1)

    airlift_parsed, timbren_parsed = run_parsing(airlift_raw, timbren_raw)

    df_airlift, df_timbren, df_all = build_dataframes(airlift_parsed, timbren_parsed)

    write_excel(df_airlift, df_timbren, df_all, OUTPUT_FILE)

    print("\nDone.")


if __name__ == "__main__":
    main()
